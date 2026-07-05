import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# Style definitions
# ============================================================
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
tier_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
tier_fills = {
    'T1': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
    'T2': PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid'),
    'T3': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
    'T4': PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'),
    'T5': PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid'),
    'T6': PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid'),
}
data_font = Font(name='Arial', size=10)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_tier_row(ws, row, cols, tier_key):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = tier_font
        cell.fill = tier_fills.get(tier_key, tier_fills['T1'])
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, cols, idx):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if idx % 2 == 0:
            cell.fill = alt_fill

# ============================================================
# Sheet 1: Retail / Semi-Pro Top 30
# ============================================================
ws1 = wb.active
ws1.title = "零售工具 Top 30"

headers1 = ['排名', '工具名称', '网站', '类型', '推测月交易量', '用户量/社区规模',
            '交易所关系/官方背书', '交易所数量', '定价模式', '核心特色', '置信度', '备注']

ws1.append(headers1)
style_header(ws1, 1, len(headers1))

retail_data = [
    # Tier 1: Broker级
    ['TIER 1 — 交易所Broker级 (月交易量 $10B+)', '', '', '', '', '', '', '', '', '', '', ''],
    [1, 'Tiger.com', 'tiger.com', '终端+Broker', '$25B+/月', '300,000+用户',
     'Binance Broker Program成员(ND Broker), Bybit/OKX正式Broker. 官网声明Binance合作', '3(Binance/Bybit/OKX)',
     '免费(Crypto许可), 付费(股票期货)', '专业桌面终端(Win/Mac)+Web终端+移动App, Tick图/集群图/DOM, 内置风控/交易日志, 子账户管理, 最高45%手续费返还',
     '高', '7年历史, 唯一确认的交易所Broker Program成员'],
    [2, 'Pionex', 'pionex.com', '交易所+内置Bot', '$3B-10B/月', '百万级注册',
     '自身是持牌交易所(FinCEN/MAS), 聚合Binance+HTX流动性', '自身即交易所',
     '免费(零手续费)', '全球首个内置Bot交易所, 16+免费Bot(Grid/DCA/Infinity Grid等), 零手续费大部分交易对',
     '高', '高榕/顺为/真格投资'],

    # Tier 2: Bot平台大户
    ['TIER 2 — Bot平台大户 (月交易量 $1B-10B)', '', '', '', '', '', '', '', '', '', '', ''],
    [3, '3Commas', '3commas.io', 'Bot平台', '$2B-8B/月', '118,000+活跃社区, 百万级注册(自2017)',
     '16交易所API集成, 无确认的Broker关系', '16',
     '$0-$49/月', 'DCA Bot/Grid Bot/Signal Bot, QuantPilot AI策略, TradingView集成, SmartTrade智能交易',
     '中高', '2017年成立, BVI注册, 行业最知名Bot平台之一'],
    [4, 'Bitsgap', 'bitsgap.com', 'Bot+聚合平台', '$1B-5B/月', '800,000+注册, 4.7M Bot启动',
     '17+交易所API集成, $9.46B用户资管规模(自称)', '17+',
     '$23-$119/月', 'GRID/DCA/BTD/LOOP/COMBO Bot, AI助手, $203M一年Bot利润(自称), 11%月均Grid Bot收益(自称)',
     '中高', '2017年成立, 迪拜FZCO注册'],
    [5, 'Cryptohopper', 'cryptohopper.com', 'Bot平台', '$1B-5B/月', '1,151,419注册',
     '11+交易所API集成, Bloomberg/Forbes/CoinDesk报道', '11+',
     '免费-$129/月', '策略/信号/模板市场, AI交易, 社交交易, DCA/套利/做市工具',
     '中', '2017年成立, 阿姆斯特丹, 自举盈利'],
    [6, 'WunderTrading', 'wundertrading.com', 'Bot平台', '$1B+/月(自称)', '210,000+交易者, 70,000+ Bot运行中',
     '20+交易所API, Gate.io官方合作伙伴', '20+',
     '免费-$44.95/月', 'DCA/Grid/Signal/Market-Neutral Bot, 套利终端, Pump Screener, MCP AI代理支持, Forbes报道',
     '中高', '自称30天交易量$1B+'],
    [7, 'MoonTrader', 'moontrader.com', '终端(手动+算法)', '$3B+/月(自称,数据矛盾)', '4,673活跃交易者',
     'Bybit Partner Program确认(partner.bybit.com/b/moontrader), Binance/OKX affiliate级别', '3(Binance/Bybit/OKX)',
     '付费许可+Bonus Program免费', '手动+算法交易, Tick图, 7种交易算法, Telegram Bot, 跨平台(Win/Mac/Linux), Prop公司服务',
     '中', '2017年成立, 爱沙尼亚. 网站一处显示3B+一处显示300B+, 数据不一致'],

    # Tier 3: 中大体量
    ['TIER 3 — 中大体量 (月交易量 $500M-3B)', '', '', '', '', '', '', '', '', '', '', ''],
    [8, 'Coinrule', 'coinrule.com', 'Bot平台', '$500M-3B/月', '1,395,960策略创建, 移动App',
     '20+交易所+DeFi链API集成, 无Broker关系. TechCrunch/Forbes报道', '20+',
     '免费-$449/月', '350+预置Bot策略, 无代码规则构建器, TradingView集成, DeFi链上交易(Base/Arbitrum/Hyperliquid)',
     '中低', 'Y Combinator S21投资, 伦敦'],
    [9, 'Tree of Alpha', 'treeofalpha.com', '新闻聚合+终端', '$1B-3B/月', '极小用户群(鲸鱼/机构级)',
     'Binance/Bybit/OKX API集成, 无Broker关系', '3',
     '$500-$2,500/月', '最快Crypto新闻聚合器(1,150+源/2,300+ Twitter), 模块化终端, 多账户管理(100+账户), $10B+累计交易量',
     '中', '定价筛出大资金交易者, $TREE代币(Ethereum)'],
    [10, 'PrimeXBT/Covesting', 'primexbt.com', 'Broker+跟单', '$500M-2B/月', '"百万级"(含传统资产), 150+国家',
     '自有Broker平台, MT5支持, COV代币', '自有平台',
     '0%起+杠杆费', 'Crypto+TradFi混合Broker, Covesting跟单模块, 最高1000x杠杆(CFD), BTC抵押交易黄金/外汇/股票',
     '中', '2018年成立, 有牌照监管'],
    [11, 'Altrady', 'altrady.com', '交易终端', '$500M-2B/月', '130,000+注册, Trustpilot 4.8/5(400+评价)',
     '17+交易所API集成, 明确声明"非Broker/不执行交易", KOL合作(Michaël van de Poppe)', '17+',
     '免费-$58/月', 'Signal Bot/Grid Bot, 多图表, TA Scanner(100+指标), 风险计算器, Paper Trading, 回测, 每周4次直播',
     '中低', '用户口碑最好的全功能终端'],
    [12, 'Freqtrade', 'freqtrade.io', '开源Bot(FOSS)', '$500M-2B/月', '50,700 GitHub stars, 10,600 forks',
     '12+交易所(通过CCXT可扩展更多), 社区驱动', '12+',
     '免费开源(GPL-3.0)', '最大的开源Crypto Bot, Python, 完整回测, FreqAI ML模块, Hyperopt策略优化, Dry-run模式',
     '中', '极其活跃的开发(31,941 commits), 最新v2026.4'],
    [13, 'Zignaly', 'zignaly.com', '跟单/利润分享', '$500M-1.5B/月', '500,000+用户',
     'Binance Labs投资, Z-Score绩效评分', '多交易所',
     '利润分享模式(盈利才付费)', '非托管跟单, 利润分享模式, $7B+累计交易(自称)',
     '中', 'Binance Labs投资是重要背书'],
    [14, 'BitFrog', 'bitfrog.io', 'Broker(返佣+跟单)', '$500M-2B/月', '未公开',
     'OKX/Binance/Bitget/Bybit/Hyperliquid展示为Partners, Broker架构(日结返佣)', '5+',
     '返佣模式', '交易返佣+Hyperliquid跟单+融资交易, 信号聚合引擎, 多渠道预警(App/电话/TG/飞书/WeCom)',
     '中', '中文市场为主, ChainUp技术合作'],
    [15, 'CScalp', 'cscalp.com', 'Scalping终端', '$500M-2B/月', '12,000 DAU, Discord 16,238, 50+国家',
     'Bitget官方合作公告, Bybit联合直播(co-marketing), 9+交易所', '9+',
     '免费', '专业Scalping终端, DOM/集群/交易带, 2008年底层(Privod Bondar), 免费模式通过交易所返佣盈利',
     '中高', '用户数据有第三方PR验证, 四语社区(RU/EN/ES/CN)'],
    [16, 'HaasOnline', 'haasonline.com', '高级Bot平台', '$500M-1B/月', '专业用户群',
     '多交易所, Kraken Futures合作', '多个',
     '$7.5-$41.5/月', '$6.4B累计交易/82M订单/13.8M回测, HaasScript自研脚本语言, TradeServer Cloud+Enterprise自托管',
     '中', '2014年成立, 行业最老Bot平台之一'],

    # Tier 4: 中小体量
    ['TIER 4 — 中小体量 (月交易量 $100M-500M)', '', '', '', '', '', '', '', '', '', '', ''],
    [17, 'CCXT', 'ccxt.com / github.com/ccxt/ccxt', '开源库(基础设施)', '不可直接比较(间接>>$100B)', '42,600 GitHub stars, 95,739 commits',
     '110+交易所官方认证(Certified), Binance/OKX/Bybit/Gate等主动维护API兼容', '110+',
     '免费+CCXT Pro付费', '行业标准Crypto交易API库, JS/TS/Python/C#/PHP/Go, 几乎所有量化框架依赖',
     '极高', '类别特殊: 基础设施层而非终端产品'],
    [18, 'Hummingbot', 'hummingbot.org', '开源做市Bot', '$300M-1B/月', '6K+ GitHub, 4,400+独立做市商',
     'Binance/OKX/Gate/Hyperliquid/KuCoin官方赞助', '数十个CEX+DEX',
     '免费开源(Apache 2.0)', '做市/套利/XEMM策略, Condor(TG管理), Gateway(DEX中间件), MCP AI代理, Cornell学术合作',
     '中高', '$3.3B交易量(Gate.io页面显示), CoinAlpha原始开发'],
    [19, 'GoodCrypto', 'goodcrypto.app', '移动端终端+Bot', '$200M-800M/月', '400,000+下载, $5B累计交易',
     '40+ CEX/DEX, GOOD代币(50%收益分享)', '40+',
     '免费+Pro付费', '最佳移动端多交易所终端, 高级订单类型, DCA/Grid/Sniper Bot, DEX交易(MPC钱包)',
     '中', '5年+历史, GOOD代币新上线'],
    [20, 'Stoic AI', 'stoic.ai', 'AI量化Bot', '$100M-500M/月', '18,000+客户, $230M AUM',
     '6交易所(Binance/Bybit/Coinbase/KuCoin/Crypto.com等)', '6',
     '$9-$19/月+AUM费', '"口袋里的对冲基金", 预构建量化策略(Index/Meta/Fixed Income/BTC Yield), Cindicator团队(2015)',
     '中', '$9M+ R&D投入, 非DIY模式'],
    [21, 'Cornix', 'cornix.io', 'Telegram信号Bot', '$200M-500M/月', '数千(信号群体)',
     '11交易所', '11',
     '订阅制', 'Telegram信号自动化执行, 群组管理工具, DCA/Grid/TradingView Bot',
     '低', '主要服务Telegram信号提供者'],
    [22, 'Moonbot', 'moon-bot.com', 'Scalping终端', '$200M-800M/月', '未公开(CIS社群深厚)',
     '6交易所(Binance/HTX/Bybit/Gate/Bitget/Hyperliquid)', '6',
     '免费+PRO/MoonScalper付费', 'Tick图交易, 40交易对同时, 检测系统, 信托管理(UDP), 出版实体书"Scalpers"',
     '低', '2017年成立, 俄语区Scalper社群根基深'],
    [23, 'TabTrader', 'tabtrader.com', '移动端终端', '$100M-500M/月', '1,000,000+ App下载',
     '20+交易所, TTT代币', '20+',
     '免费+Pro付费', '移动优先多交易所终端, 实时图表/智能预警/套利工具, TabTrader Academy',
     '低', '高下载量但活跃交易量未知'],
    [24, 'Alertatron', 'alertatron.com', 'TradingView自动化', '$100M-300M/月', '未公开',
     '10交易所API(Bybit/Binance/OKX/Bitget/Coinbase/Bitfinex/Deribit/BitMEX/BingX/Phemex)', '10',
     '$59-$199/月', 'TradingView信号自动执行, 复杂订单链, 200人交易群组管理(Pro), <1秒执行',
     '低', ''],
    [25, 'Gunbot', 'gunbot.com', '自托管Bot', '$100M-300M/月', '自2017年, 忠实社区',
     '20+交易所(含DeFi)', '20+',
     '一次性€44-€187(终身)', '隐私优先(本地运行), 终身买断, 20+内置策略, AI策略生成, TradingView集成, DeFi Bot',
     '低', '独特的一次性买断模式'],
    [26, 'Mudrex', 'mudrex.com', '投资+Bot平台', '$100M-300M/月', '1,000,000+用户(印度为主)',
     '多交易所, YC投资', '多个',
     '免费+付费', 'Coin Sets指数投资, 算法策略市场, $3B+累计交易(自称)',
     '低', 'YC投资, 印度市场专注'],
    [27, 'Bookmap', 'bookmap.com', '订单流可视化', '$50M-200M/月', '数千(专业小众)',
     '20+交易所(Crypto+股票/期货)', '20+',
     '免费-$79/月', '行业领先订单簿热力图(40FPS), Multibook聚合5交易所, 清算指标/冰山检测/大单追踪/DOM Pro',
     '低', '唯一提供此级别订单流分析的零售工具'],
    [28, 'Quantower', 'quantower.com', '多资产终端', '$50M-200M/月(Crypto部分)', '专业用户',
     '多交易所+经纪商+数据源, B2B方案', '多个',
     '免费+付费', '多资产(Crypto+期货+股票+期权), 合成商品, 多连接同步, 期权链, Windows桌面端',
     '低', '英国/乌克兰注册, Crypto只是子集'],
    [29, 'FMZ Quant', 'fmz.com', '量化策略平台', '$50M-200M/月', '大(亚洲为主)',
     '多交易所, Gate.io官方合作伙伴', '多个',
     '免费增值', '多语言策略(JS/Python/C++/Pine Script), 策略市场(租/售), Alpha 101, 移动App',
     '低', '新加坡Inventor PTE LTD, 中文社区强'],
    [30, 'OctoBot', 'octobot.cloud / github.com/Drakkar-Software/OctoBot', '开源Bot+Cloud', '$30M-100M/月', '6,000 GitHub stars',
     '15+交易所', '15+',
     '免费开源(GPL-3.0)+Cloud', 'AI交易模式(OpenAI/Ollama), Grid/DCA, TradingView连接器, 树莓派可运行, 一键DigitalOcean部署',
     '低', '最新v2026, 122个Release'],
]

tier_rows = set()
row_idx = 2
data_idx = 0
for item in retail_data:
    ws1.append(item)
    if isinstance(item[0], str) and item[0].startswith('TIER'):
        tier_key = item[0].split(' ')[1]
        tier_map = {'1': 'T1', '2': 'T2', '3': 'T3', '4': 'T4'}
        style_tier_row(ws1, row_idx, len(headers1), tier_map.get(tier_key, 'T1'))
        ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers1))
        tier_rows.add(row_idx)
    else:
        style_data_row(ws1, row_idx, len(headers1), data_idx)
        data_idx += 1
    row_idx += 1

# Column widths for sheet 1
col_widths1 = [5, 18, 30, 18, 18, 28, 40, 12, 20, 50, 8, 40]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{row_idx - 1}"
ws1.freeze_panes = 'A2'

# ============================================================
# Sheet 2: Institutional Top 30
# ============================================================
ws2 = wb.create_sheet("机构工具 Top 30")

headers2 = ['排名', '公司名称', '网站', '类型', '推测日交易量', '关键数据',
            '交易所关系/连接', '融资/估值', '核心业务', '备注']

ws2.append(headers2)
style_header(ws2, 1, len(headers2))

inst_data = [
    ['TIER 1 — 自营做市商 (日交易量 $1B+)', '', '', '', '', '', '', '', '', ''],
    [1, 'Jump Trading/Crypto', 'jumpcrypto.com', '自营做市商/HFT', '$5-15B/天', '全球最大HFT之一, 1999年成立',
     '所有主要CEX, Solana/Wormhole生态', '自有资金(数十亿)', '自营量化交易, 做市, DeFi基础设施, 跨所套利', '2023年缩减Crypto业务后部分恢复'],
    [2, 'Wintermute', 'wintermute.com', '做市商+OTC', '$5-10B/天', '50+交易所(CEX+DEX)做市',
     '所有主要CEX+DEX, 数百Token项目流动性合作', '$75M / $750M估值(Lightspeed)', '算法做市, OTC(现货/期权/远期/CFD), DeFi, Ventures', '最大的Crypto原生做市商之一'],
    [3, 'Flow Traders', 'flowtraders.com', '做市商', '$2-5B/天(Crypto)', '上市公司(Euronext FLOW.AS), ~€670M营收(2024)',
     '主要Crypto交易所, BTC/ETH现货ETF做市', '公开上市', 'ETF做市(全球领先), Crypto现货/ETF做市, 自营交易', 'Crypto ETF生态核心参与者'],
    [4, 'Cumberland (DRW)', 'cumberland.io', '做市商+OTC', '$1-5B/天', 'DRW旗下, 2014年入场Crypto',
     '主要交易所, Goldman Sachs客户(APAC数字资产主管引用)', 'DRW自有资金(数十亿)', 'OTC深度流动性(现货/期权/期货/NDF), 自营做市', '芝加哥老牌Prop Trading出身'],
    [5, 'Virtu Financial', 'virtu.com', '做市商', '$1-3B/天(Crypto)', '上市公司(NASDAQ VIRT), ~$21亿营收',
     '主要Crypto交易所', '公开上市', '全球最大电子做市商之一, Crypto做市+机构执行', 'Crypto占整体业务比例较小但在增长'],
    [6, 'GSR Markets', 'gsr.io', '做市商+OTC+顾问', '$1-3B/天', '2013年成立, "2025年度Crypto做市商"',
     'UK FCA/Singapore MAS牌照, Chainlink/Polygon合作', '自有资金+$57M收购', '做市, OTC, 衍生品, 结构化产品, 顾问, 资管, ETF(BESO)', '最老的Crypto交易公司之一, 收购Autonomous+Architech'],
    [7, 'B2C2 (SBI)', 'b2c2.com', '做市商+OTC', '$1-3B/天', 'SBI集团收购(2020), 首个MiCA授权OTC LP(2026.5)',
     '银行/经纪商/交易所/Fintech全球客户', 'SBI集团子公司', 'OTC流动性(现货/期权/期货/远期/CFD/NDF), Penny稳定币互换', '伦敦总部, 7个全球办公室'],

    ['TIER 2 — Prime Broker (日路由量 $1B+)', '', '', '', '', '', '', '', '', ''],
    [8, 'FalconX', 'falconx.io', 'Prime Broker', '$3-8B/天', '$2.5T+累计, 机构期权Block Volume #1, 收购21Shares',
     'Binance/OKX/Bybit/Coinbase/Kraken等DMA, CFTC注册', '$417M / $8B估值(Tiger Global/Accel)', 'Trading/Financing/Custody/DMA/ETF Solutions/电子期权', '350+员工, 7个全球办公室, 2025收购21Shares'],
    [9, 'Hidden Road (Ripple)', 'hiddenroad.com', 'Prime Broker', '$3-10B/天(多资产)', '$3T+年清算, 30M+日交易笔, 300+机构客户',
     '多资产(Crypto+FX+固收), 2025被Ripple $12.5亿收购', '收购价$1.25B', '无冲突PB/清算/融资, Route28 OTC产品, 衍生品清算', '当时Crypto最大收购案'],
    [10, 'Galaxy Digital', 'galaxy.com', 'PB+自营+资管', '$500M-2B/天', '上市公司(NASDAQ), $5B+ AUM, Mike Novogratz',
     '所有主要交易所, GalaxyOne Prime服务', '公开上市(NASDAQ)', 'OTC/衍生品/Prime/投行/资管/Staking/代币化/AI数据中心', '收购GK8(托管)/Caspian(交易基础设施)'],
    [11, 'Amber Group', 'ambergroup.io', 'OTC+资管', '$500M-1B/天', '$1T+累计, 1,000+机构客户',
     '摩根士丹利校友创立, 港新总部', '$300M / $3B估值(淡马锡/分布式)', 'Algo执行/OTC/结构化产品/资管/DeFi', '2017年成立'],
    [12, 'Matrixport', 'matrixport.com', '交易+结构化产品', '$300M-1B/天', '4M+用户, $50B+累计交易(自称)',
     '吴忌寒创立(Bitmain联合创始人), 亚洲最大', '$100M+ / $1.5B估值(Paradigm/Dragonfly)', 'Trading/结构化产品/双币投资/借贷/托管', 'Bitmain分拆, 亚洲机构+高净值为主'],

    ['TIER 3 — 交易基础设施 (促成日交易量 $500M+)', '', '', '', '', '', '', '', '', ''],
    [13, 'Talos', 'talos.com', '交易基础设施(OMS/EMS)', '促成$500M-2B/天', '连接~60交易场所, 100+总集成, 收购Coin Metrics',
     '35+交易所(Binance/OKX/Bybit/Bitget/Coinbase等)+OTC流+Prime Broker+托管', '$153M / $1.3B估值(a16z/花旗/BNY Mellon/Wells Fargo)', '"Crypto的Bloomberg Terminal", 全交易生命周期, 白标方案, FIX连接', '银行/资管/对冲基金/ETF发行商/OTC/Prop公司客户'],
    [14, 'Paradigm', 'paradigm.co', '衍生品RFQ网络', '数十亿/天(期权为主)', '1,000+对手方, 50+做市商, 120+产品, 20+ Prime Dealers',
     'Deribit/Bybit/DeFi协议结算', '~$50M+(Jump Crypto/Dragonfly)', '非托管RFQ平台, 机构期权/永续/期货/多腿策略', '主导Crypto期权Block交易'],
    [15, 'Fireblocks', 'fireblocks.com', '托管+基础设施', '$6T+累计转移', '1,800+机构客户(BNY Mellon/ANZ/Revolut/Galaxy)',
     'MPC钱包/Exchange Link/DeFi接入', '$1.1B / $8B估值(Sequoia/Coatue/BNY)', 'MPC托管/转账/交易结算/代币化/支付/Staking', '间接促成交易: 每笔通过Fireblocks的交易结算=交易所成交量'],
    [16, 'SFOX', 'sfox.com', '机构Prime Dealer', '$500M-1B/天', '聚合30+流动性源, 80+市场, 15+算法',
     'FinCEN MSB+MTL牌照, SAFE Trust托管', '~$75M+', '流动性聚合/智能路由/托管/Staking/Prime/白标(sFOX Connect)', ''],
    [17, 'Copper.co', 'copper.co', '托管+ClearLoop', '促成$数十亿/天', '$10B+托管, ClearLoop=免预充值交易所交易',
     'Binance/OKX/Bybit/Deribit ClearLoop连接', '$150M+ / $2B估值(Barclays/Tiger Global)', '托管/ClearLoop结算/DeFi接入', 'ClearLoop直接促成交易所成交量(降低对手方风险)'],

    ['TIER 4 — ECN/SOR/专业平台 (日交易量 $50M-500M)', '', '', '', '', '', '', '', '', ''],
    [18, 'Finery Markets', 'finerymarkets.com', 'Crypto ECN', '$100M-500M/天', 'CNBC 300 Fintech 2025, Deloitte Fast 50, ISO 27001',
     'IMC/ZebPay/Wintermute/Flow Traders/BitGo客户', '~$5.5M', '非托管ECN(FM Marketplace), Trading SaaS(FM Liquidity Match), 白标, 数据分析(FM Pulse)', '200+ Crypto+法币交易对'],
    [19, 'CoinRoutes', 'coinroutes.com', 'SOR/算法交易', '$50M-200M/天', '42 CEX/DEX/OTC场所, Gate.io官方合作',
     'DCG投资', '~$7M(DCG)', '智能订单路由(TWAP/VWAP等), 实时分析, 机构执行', ''],
    [20, 'WOO Network', 'woo.org', '流动性聚合+DEX', '$100M-500M/天', '连接Binance/OKX/Bybit/KuCoin/Kraken',
     '交易所流动性聚合+WOOFi DEX+WOOFi Pro', '~$50M+', '流动性聚合/WOOFi(DEX跨链互换)/WOOFi Pro(无Gas永续DEX)/Starchild(AI)', '从流动性聚合器转型DEX+AI'],
    [21, 'Floating Point Group', 'floatingpoint.group', '机构Broker', '$100M-500M/天', 'DMA聚合, SOC 2认证',
     '主要交易所DMA', '~$50M(10T Holdings)', '执行/结算/托管, 单API简化机构交易', '可能已缩减运营(域名状态不确定)'],

    ['TIER 5 — 托管/银行/资管 (间接促成交易量)', '', '', '', '', '', '', '', '', ''],
    [22, 'BitGo Prime', 'bitgo.com', '托管+Prime', '中等直接量', '$100B+托管, ~20%链上BTC交易',
     'OTC交易+结算', '$470M / $1.75B估值', '机构托管/OTC/借贷/Staking/NFT', '差点被Galaxy收购(2022)'],
    [23, 'Anchorage Digital', 'anchorage.com', 'Crypto银行', '中等', '唯一OCC联邦特许Crypto银行, $10B+托管',
     'a16z/Goldman Sachs/Visa/KKR投资', '$487M / $3B估值', '托管/Staking/交易/借贷/治理', ''],
    [24, '1Token', '1token.trade', '机构资管软件', '促成$数十亿', '$20B+ AUM客户, SOC 2, Gate.io官方合作',
     '机构投资管理平台', '未公开', '数字资产投资管理软件, 对冲基金/资管/基金管理', '新加坡Inventor PTE LTD(与FMZ Quant关联)'],
    [25, 'Margin.de', 'margin.de', '机构HFT引擎', '$300M+/月', '已从零售转型机构, Rust底层',
     '对冲基金客户', '未公开', 'Rust HFT交易引擎, 机构专用', '不再服务零售'],
    [26, 'Keyrock', 'keyrock.eu', '做市商', '$数亿/天', '欧洲合规做市商, MiCA牌照',
     'Token流动性合作', '$72M(SBI)', '算法做市/OTC/Treasury管理', '布鲁塞尔总部, 欧洲合规先行者'],
    [27, 'Cobo', 'cobo.com', '托管/钱包', '低直接量', '$2B+托管, 亚洲主导',
     'DST Global投资', '~$50M', 'MPC钱包/智能合约钱包/交易所钱包/托管', ''],
    [28, 'Hex Trust', 'hextrust.com', '亚洲托管', '低直接量', '香港/新加坡/迪拜牌照',
     'Ripple/Animoca Brands/Morgan Creek投资', '~$88M', '托管/Staking/DeFi/经纪', ''],
    [29, 'DV Chain', 'dvchain.co', 'OTC Desk', '$数亿/天', '芝加哥, 机构OTC流动性',
     'Talos网络OTC提供商', '未公开', 'OTC Streaming/RFQ流动性, 自营做市', ''],
    [30, 'Enigma Securities', 'enigma-securities.io', 'OTC+经纪', '$数亿/天', '伦敦FCA监管',
     'Talos网络OTC提供商', '未公开', '机构OTC流动性/经纪/结算', ''],
]

tier_rows2 = set()
row_idx2 = 2
data_idx2 = 0
for item in inst_data:
    ws2.append(item)
    if isinstance(item[0], str) and item[0].startswith('TIER'):
        tier_key = item[0].split(' ')[1]
        tier_map = {'1': 'T1', '2': 'T2', '3': 'T3', '4': 'T4', '5': 'T5'}
        style_tier_row(ws2, row_idx2, len(headers2), tier_map.get(tier_key, 'T1'))
        ws2.merge_cells(start_row=row_idx2, start_column=1, end_row=row_idx2, end_column=len(headers2))
        tier_rows2.add(row_idx2)
    else:
        style_data_row(ws2, row_idx2, len(headers2), data_idx2)
        data_idx2 += 1
    row_idx2 += 1

col_widths2 = [5, 22, 28, 22, 18, 40, 40, 25, 50, 40]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{row_idx2 - 1}"
ws2.freeze_panes = 'A2'

# ============================================================
# Sheet 3: Exchange Endorsement
# ============================================================
ws3 = wb.create_sheet("交易所背书等级")

headers3 = ['等级', '工具', '背书类型', '具体证据']
ws3.append(headers3)
style_header(ws3, 1, len(headers3))

endorsement_data = [
    ['S', 'CCXT', '交易所主动认证(Certified)', '110+交易所主动维护API兼容, Binance/OKX/Bybit/Gate/KuCoin等官方标记"CCXT Certified"'],
    ['A', 'Tiger.com', 'Binance Broker Program(ND Broker)', '官网声明Binance Broker合作+链接, ND Broker架构(子账户/手续费返还/自定义杠杆), Bybit/OKX正式Broker'],
    ['A-', 'Pionex', '自身是持牌交易所', 'FinCEN/MAS牌照, 聚合Binance+HTX流动性, 不是"背书"而是"自身就是交易所"'],
    ['B+', 'Hummingbot', '交易所官方赞助', 'Binance/OKX/Gate.io/Hyperliquid/KuCoin赞助(Sponsor), Gate.io官方合作伙伴页列出'],
    ['B', 'CScalp', '交易所官方合作公告', 'Bitget官方发布合作公告(2024.5), Bybit联合直播(Bybit x CScalp), Referral返佣'],
    ['B', 'BitFrog', 'Broker架构+Partner展示', '日结返佣/API Algo返佣(匹配Broker架构), OKX/Binance/Bitget/Bybit/Hyperliquid Logo展示为Partners, 但缺交易所端确认'],
    ['B', 'WunderTrading', 'Gate.io官方合作伙伴', 'Gate.io Broker页面公开列出为合作伙伴'],
    ['B', 'Bitsgap', 'Gate.io官方合作伙伴', 'Gate.io Broker页面公开列出(500,000+活跃用户)'],
    ['B', '3Commas', 'Gate.io官方合作伙伴', 'Gate.io Broker页面公开列出(19+交易所)'],
    ['B-', 'MoonTrader', 'Bybit Partner确认', 'partner.bybit.com/b/moontrader链接格式=正式Partner注册, 费率返还(最高30%)'],
    ['B-', 'Zignaly', 'Binance Labs投资', 'Binance Labs投资=Binance生态认可, 但不等于Broker关系'],
    ['C', 'Coinrule', '仅API集成', 'Y Combinator投资但无交易所官方背书, 20+交易所API连接'],
    ['C', 'Altrady', '仅API集成', '明确声明"非Broker/不执行交易/纯软件提供商", 17+交易所API'],
    ['C', 'Cryptohopper', '仅API集成', '1.15M注册但无交易所官方背书, Bloomberg/Forbes报道≠交易所背书'],
    ['C', 'Alertatron/Cornix/TabTrader等', '仅API集成', '功能性API连接, 无正式商业合作关系'],
]

r = 2
for i, row_data in enumerate(endorsement_data):
    ws3.append(row_data)
    style_data_row(ws3, r, len(headers3), i)
    # Color code the grade
    grade_cell = ws3.cell(row=r, column=1)
    grade_colors = {
        'S': 'FF0000', 'A': 'ED7D31', 'A-': 'FFC000',
        'B+': '70AD47', 'B': '5B9BD5', 'B-': '8DB4E2', 'C': 'D9D9D9'
    }
    grade = row_data[0]
    if grade in grade_colors:
        grade_cell.fill = PatternFill(start_color=grade_colors[grade], end_color=grade_colors[grade], fill_type='solid')
        if grade in ('S', 'A'):
            grade_cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    r += 1

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 80
ws3.freeze_panes = 'A2'

# ============================================================
# Sheet 4: Volume Comparison
# ============================================================
ws4 = wb.create_sheet("量级对比")

headers4 = ['类别', '头部日交易量', '典型代表', '说明']
ws4.append(headers4)
style_header(ws4, 1, len(headers4))

comparison = [
    ['自营做市商', '$1-15B/天', 'Jump, Wintermute, Flow Traders, Cumberland', '交易所上最大的交易量来源, 自有资金持续报价和套利'],
    ['Prime Broker', '$1-10B/天', 'FalconX, Hidden Road, Galaxy', '路由机构客户订单到交易所, 单客户可达$100M+/天'],
    ['交易基础设施', '促成$500M-2B/天', 'Talos, Paradigm, SFOX', '软件平台, 机构通过它交易, 每笔订单=交易所成交量'],
    ['零售Broker(最大)', '$500M-1B/天', 'Tiger.com, Pionex', '交易所Broker Program成员, 路由零售用户交易'],
    ['零售Bot平台(最大)', '$30M-300M/天', '3Commas, Bitsgap, Cryptohopper', 'API连接, 用户自有账户交易, 平台不路由订单'],
    ['零售终端(最大)', '$10M-100M/天', 'Altrady, CScalp, TabTrader', 'UI工具, 不直接产生额外交易量'],
    ['开源库/框架', '不可直接比较', 'CCXT, Freqtrade, Hummingbot', '间接促成的交易量远超所有终端产品总和'],
]

for i, row_data in enumerate(comparison):
    ws4.append(row_data)
    style_data_row(ws4, i + 2, len(headers4), i)

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 45
ws4.column_dimensions['D'].width = 60
ws4.freeze_panes = 'A2'

# ============================================================
# Save
# ============================================================
output_path = '/root/projects/Crypto_Trading_Tools_Top30_Report.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
